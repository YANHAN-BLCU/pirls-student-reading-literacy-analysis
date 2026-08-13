import os
from openpyxl import load_workbook

# 定义值的组合
value_combinations = [
    {9134, 9135, 9130, 9133},
    {643, 643100},
    {710, 7106},
    {784, 7842, 7841},
    {956, 957}
]

# 定义文件夹路径
folder_path = 'E:\北语\大创项目\COMB\A5R5'

# 存储每个组合对应的文件列表
file_groups = {frozenset(comb): [] for comb in value_combinations}

# 遍历文件夹中的所有文件
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file.endswith('.xlsx'):
            file_path = os.path.join(root, file)
            try:
                # 加载工作簿
                wb = load_workbook(file_path)
                # 获取第一个工作表
                ws = wb.active
                # 读取 A2 单元格的值
                value = ws['A2'].value
                # 检查值是否在某个组合中
                for comb in value_combinations:
                    if value in comb:
                        file_groups[frozenset(comb)].append(file_path)
                        break
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

# 合并每个组合中的文件
for comb, files in file_groups.items():
    if len(files) > 1:
        # 第一个文件作为主文件
        main_file = load_workbook(files[0])
        main_ws = main_file.active
        # 遍历其他文件
        for file in files[1:]:
            wb = load_workbook(file)
            ws = wb.active
            # 从第二行开始复制数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                main_ws.append(row)
        # 保存合并后的文件
        main_file.save(files[0])
        print(f"Merged files for combination {comb} into {files[0]}")

import os
print(os.getcwd())  # 打印当前工作目录